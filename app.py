import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta, time
import io
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure photo directory exists
os.makedirs("uploads", exist_ok=True)

# Initialize Database Connection
conn = sqlite3.connect("inventory.db", check_same_thread=False)
cursor = conn.cursor()

# --- DATABASE SETUP ---
cursor.execute("""
CREATE TABLE IF NOT EXISTS master_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_name TEXT UNIQUE,
    category TEXT,
    unit TEXT,
    current_stock REAL,
    min_threshold REAL
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT,
    item_name TEXT,
    type TEXT,
    quantity REAL,
    user_role TEXT,
    driver_details TEXT,
    issued_to TEXT,
    project_name TEXT,
    purpose TEXT,
    remarks TEXT,
    photo_path TEXT,
    edit_status TEXT DEFAULT 'NORMAL'
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS edit_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    transaction_id INTEGER,
    requested_by TEXT,
    reason TEXT,
    original_data TEXT,
    proposed_data TEXT,
    status TEXT DEFAULT 'PENDING',
    review_remarks TEXT,
    timestamp TEXT,
    FOREIGN KEY(transaction_id) REFERENCES transactions(id)
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    target_user TEXT,
    target_role TEXT,
    message TEXT,
    is_read INTEGER DEFAULT 0,
    timestamp TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    password TEXT,
    role TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS reminders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_name TEXT,
    title TEXT,
    due_date TEXT,
    priority TEXT,
    status TEXT DEFAULT 'PENDING',
    timestamp TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS schedules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_name TEXT,
    title TEXT,
    event_date TEXT,
    start_time TEXT,
    end_time TEXT,
    location_details TEXT,
    notes TEXT,
    timestamp TEXT
)
""")

# NEW TABLE: Physical Inventory Audits & Reconciliation
cursor.execute("""
CREATE TABLE IF NOT EXISTS physical_inventory_counts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    audit_date TEXT,
    item_name TEXT,
    system_stock REAL,
    physical_stock REAL,
    variance REAL,
    unit TEXT,
    counted_by TEXT,
    remarks TEXT,
    sync_status TEXT DEFAULT 'UNSYNCED',
    timestamp TEXT
)
""")

# Schema Safe Migrations
for col, col_type in [
    ("driver_details", "TEXT"),
    ("issued_to", "TEXT"),
    ("project_name", "TEXT"),
    ("purpose", "TEXT"),
    ("photo_path", "TEXT"),
    ("edit_status", "TEXT DEFAULT 'NORMAL'")
]:
    try:
        cursor.execute(f"ALTER TABLE transactions ADD COLUMN {col} {col_type}")
    except sqlite3.OperationalError:
        pass

# Seed Default Accounts
cursor.execute("SELECT COUNT(*) FROM users")
if cursor.fetchone()[0] == 0:
    cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", 
                   ("admin", "admin123", "Head Office"))
    cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", 
                   ("supervisor", "super123", "Materials Supervisor"))

# Seed Expanded Inventory List
cursor.execute("SELECT COUNT(*) FROM master_items")
if cursor.fetchone()[0] == 0:
    sample_items = [
        ("Diesel Fuel", "1. Fuel & Oils", "Liters", 1200, 500),
        ("Gasoline", "1. Fuel & Oils", "Liters", 200, 50),
        ("Engine Oil #10", "1. Fuel & Oils", "Liters", 40, 20),
        ("Engine Oil #40", "1. Fuel & Oils", "Liters", 40, 20),
        ("Hydraulic Oil #68", "1. Fuel & Oils", "Pails (20L)", 8, 3),
        ("Chassis Grease", "1. Fuel & Oils", "Cans (1kg)", 15, 5),
        ("Tonner Cement", "2. Construction Materials", "Bags (1-Ton)", 15, 5),
        ("Portland Cement", "2. Construction Materials", "Bags (40kg)", 250, 100),
        ("Plywood 1/2 (4x8)", "2. Construction Materials", "Sheets", 80, 25),
        ("Plywood 3/4 (4x8)", "2. Construction Materials", "Sheets", 50, 15),
        ("Coco Lumber 2x2x12", "2. Construction Materials", "Board Feet", 300, 100),
        ("Rebar 10mm x 6m", "3. Steel / Rebar", "Pcs", 350, 100),
        ("Rebar 12mm x 6m", "3. Steel / Rebar", "Pcs", 250, 80),
        ("Rebar 16mm x 6m", "3. Steel / Rebar", "Pcs", 150, 50),
        ("G.I. Tie Wire #16", "3. Steel / Rebar", "Kilos", 50, 15),
        ("CWN #1-1/2 (Common Nails)", "4A. Nails & Fasteners", "Kilos", 30, 10),
        ("CWN #2 (2 Common Nails)", "4A. Nails & Fasteners", "Kilos", 45, 20),
        ("CWN #3 (3 Common Nails)", "4A. Nails & Fasteners", "Kilos", 40, 15),
        ("CWN #4 (4 Common Nails)", "4A. Nails & Fasteners", "Kilos", 35, 10),
        ("Concrete Nails 3", "4A. Nails & Fasteners", "Kilos", 20, 5),
        ("Cutting Disc 4", "4B. Cutting & Grinding Consumables", "Pcs", 120, 50),
        ("Grinding Disc 4", "4B. Cutting & Grinding Consumables", "Pcs", 60, 20),
        ("Diamond Cutter Blade 14", "4B. Cutting & Grinding Consumables", "Pcs", 4, 2),
        ("Welding Rod 6011", "4C. Welding Supplies & PPE", "Kilos", 30, 15),
        ("Welding Rod 6013", "4C. Welding Supplies & PPE", "Kilos", 50, 20),
        ("Safety Helmets (Hard Hats)", "4C. Welding Supplies & PPE", "Pcs", 25, 10),
        ("Cotton Gloves (Pair)", "4C. Welding Supplies & PPE", "Pairs", 100, 30),
        ("Chalk Stone (Marking)", "4D. General Site Supplies", "Boxes", 12, 5),
        ("Nylon Rope 1/2", "4D. General Site Supplies", "Meters", 100, 30)
    ]
    cursor.executemany("""
    INSERT INTO master_items (item_name, category, unit, current_stock, min_threshold)
    VALUES (?, ?, ?, ?, ?)
    """, sample_items)

conn.commit()


# --- HELPER: NOTIFICATION CREATOR ---
def create_notification(message, target_user=None, target_role=None):
    cursor.execute("""
        INSERT INTO notifications (target_user, target_role, message, is_read, timestamp)
        VALUES (?, ?, ?, 0, ?)
    """, (target_user, target_role, message, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()


# --- MULTI-SHEET EXCEL GENERATOR ---
def generate_excel_report(user_name, selected_date, df_daily_tx, df_current_stock):
    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws_master = wb.active
    ws_master.title = "Master Activity Log"
    ws_master.views.sheetView[0].showGridLines = True

    ws_master.merge_cells("A1:I1")
    title1 = ws_master["A1"]
    title1.value = "MASTER SITE INVENTORY ACTIVITY LOG"
    title1.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title1.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title1.alignment = Alignment(horizontal="center", vertical="center")
    ws_master.row_dimensions[1].height = 30

    ws_master["A3"] = f"Report Date: {selected_date}"
    ws_master["A3"].font = Font(bold=True)
    ws_master["A4"] = f"Supervisor: {user_name}"
    ws_master["A4"].font = Font(bold=True)

    headers_master = [
        "Time", "Item Name", "Type", "Quantity", "Issued To / Recipient", 
        "Driver / Delivery", "Project Name", "Purpose / Equipment", "Logged By"
    ]
    for col_num, h_title in enumerate(headers_master, 1):
        cell = ws_master.cell(row=6, column=col_num, value=h_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 7
    if not df_daily_tx.empty:
        for _, r in df_daily_tx.iterrows():
            ws_master.cell(row=row_idx, column=1, value=str(r.get('timestamp', '')))
            ws_master.cell(row=row_idx, column=2, value=str(r.get('item_name', '')))
            
            t_cell = ws_master.cell(row=row_idx, column=3, value=str(r.get('type', '')))
            t_cell.alignment = Alignment(horizontal="center")
            t_cell.font = Font(bold=True, color="006100" if r.get('type') == 'IN' else "9C0006")

            q_cell = ws_master.cell(row=row_idx, column=4, value=float(r.get('quantity', 0)))
            q_cell.number_format = "#,##0.00"
            
            ws_master.cell(row=row_idx, column=5, value=str(r.get('issued_to', '-')))
            ws_master.cell(row=row_idx, column=6, value=str(r.get('driver_details', '-')))
            ws_master.cell(row=row_idx, column=7, value=str(r.get('project_name', '-')))
            ws_master.cell(row=row_idx, column=8, value=str(r.get('purpose', r.get('remarks', '-'))))
            ws_master.cell(row=row_idx, column=9, value=str(r.get('user_role', '')))

            for c in range(1, 10):
                ws_master.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
    else:
        ws_master.cell(row=7, column=1, value="No activity recorded for this date.")

    ws_in = wb.create_sheet(title="Stock IN Entries")
    ws_in.views.sheetView[0].showGridLines = True
    ws_in.merge_cells("A1:F1")
    title2 = ws_in["A1"]
    title2.value = "DAILY STOCK IN / RECEIVING REPORT"
    title2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title2.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    title2.alignment = Alignment(horizontal="center", vertical="center")
    ws_in.row_dimensions[1].height = 30

    headers_in = ["Time", "Item Name", "Qty Received", "Driver / Delivery Details", "Logged By", "General Remarks"]
    for col_num, h_title in enumerate(headers_in, 1):
        cell = ws_in.cell(row=4, column=col_num, value=h_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df_in = df_daily_tx[df_daily_tx['type'] == 'IN'] if not df_daily_tx.empty else pd.DataFrame()
    row_idx = 5
    if not df_in.empty:
        for _, r in df_in.iterrows():
            ws_in.cell(row=row_idx, column=1, value=str(r.get('timestamp', '')))
            ws_in.cell(row=row_idx, column=2, value=str(r.get('item_name', '')))
            q_cell = ws_in.cell(row=row_idx, column=3, value=float(r.get('quantity', 0)))
            q_cell.number_format = "#,##0.00"
            q_cell.font = Font(bold=True, color="006100")
            
            ws_in.cell(row=row_idx, column=4, value=str(r.get('driver_details', '-')))
            ws_in.cell(row=row_idx, column=5, value=str(r.get('user_role', '')))
            ws_in.cell(row=row_idx, column=6, value=str(r.get('remarks', '-')))

            for c in range(1, 7):
                ws_in.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
    else:
        ws_in.cell(row=5, column=1, value="No Stock IN entries for this date.")

    ws_out = wb.create_sheet(title="Stock OUT Entries")
    ws_out.views.sheetView[0].showGridLines = True
    ws_out.merge_cells("A1:H1")
    title3 = ws_out["A1"]
    title3.value = "DAILY STOCK OUT / ISSUANCE REPORT"
    title3.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title3.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    title3.alignment = Alignment(horizontal="center", vertical="center")
    ws_out.row_dimensions[1].height = 30

    headers_out = ["Time", "Item Name", "Qty Issued", "Issued To", "Driver / Transport", "Project Name", "Purpose / Usage", "Logged By"]
    for col_num, h_title in enumerate(headers_out, 1):
        cell = ws_out.cell(row=4, column=col_num, value=h_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="833C0C", end_color="833C0C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    df_out = df_daily_tx[df_daily_tx['type'] == 'OUT'] if not df_daily_tx.empty else pd.DataFrame()
    row_idx = 5
    if not df_out.empty:
        for _, r in df_out.iterrows():
            ws_out.cell(row=row_idx, column=1, value=str(r.get('timestamp', '')))
            ws_out.cell(row=row_idx, column=2, value=str(r.get('item_name', '')))
            q_cell = ws_out.cell(row=row_idx, column=3, value=float(r.get('quantity', 0)))
            q_cell.number_format = "#,##0.00"
            q_cell.font = Font(bold=True, color="9C0006")
            
            ws_out.cell(row=row_idx, column=4, value=str(r.get('issued_to', '-')))
            ws_out.cell(row=row_idx, column=5, value=str(r.get('driver_details', '-')))
            ws_out.cell(row=row_idx, column=6, value=str(r.get('project_name', '-')))
            ws_out.cell(row=row_idx, column=7, value=str(r.get('purpose', r.get('remarks', '-'))))
            ws_out.cell(row=row_idx, column=8, value=str(r.get('user_role', '')))

            for c in range(1, 9):
                ws_out.cell(row=row_idx, column=c).border = thin_border
            row_idx += 1
    else:
        ws_out.cell(row=5, column=1, value="No Stock OUT entries for this date.")

    ws_cat = wb.create_sheet(title="Categorized Stock Balance")
    ws_cat.views.sheetView[0].showGridLines = True
    ws_cat.merge_cells("A1:D1")
    title4 = ws_cat["A1"]
    title4.value = "SITE INVENTORY BALANCE (BY CATEGORY)"
    title4.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title4.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title4.alignment = Alignment(horizontal="center", vertical="center")
    ws_cat.row_dimensions[1].height = 30

    c_row = 3
    if not df_current_stock.empty:
        for cat in sorted(df_current_stock['category'].unique()):
            ws_cat.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=4)
            cat_cell = ws_cat.cell(row=c_row, column=1, value=str(cat).upper())
            cat_cell.font = Font(bold=True, color="FFFFFF")
            cat_cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            c_row += 1

            sub_headers = ["Item Name", "Current Stock", "Min Alert Threshold", "Unit"]
            for col_num, h_title in enumerate(sub_headers, 1):
                cell = ws_cat.cell(row=c_row, column=col_num, value=h_title)
                cell.font = Font(bold=True, color="333333")
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            c_row += 1
            for _, r in df_current_stock[df_current_stock['category'] == cat].iterrows():
                ws_cat.cell(row=c_row, column=1, value=str(r['item_name']))
                
                stk_cell = ws_cat.cell(row=c_row, column=2, value=float(r['current_stock']))
                stk_cell.number_format = "#,##0.00"

                thresh_cell = ws_cat.cell(row=c_row, column=3, value=float(r['min_threshold']))
                thresh_cell.number_format = "#,##0.00"

                ws_cat.cell(row=c_row, column=4, value=str(r['unit']))

                if float(r['current_stock']) <= float(r['min_threshold']):
                    stk_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    stk_cell.font = Font(color="9C0006", bold=True)

                for col_num in range(1, 5):
                    ws_cat.cell(row=c_row, column=col_num).border = thin_border
                c_row += 1
            c_row += 1

    for sheet in [ws_master, ws_in, ws_out, ws_cat]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row != 1 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- SESSION STATE INITIALIZATION ---
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "username" not in st.session_state:
    st.session_state["username"] = None
if "user_role" not in st.session_state:
    st.session_state["user_role"] = None


# --- LOGIN SCREEN ---
if not st.session_state["logged_in"]:
    st.title("🏗️ Construction Site Inventory")
    st.subheader("Sign In")

    input_username = st.text_input("Username")
    input_password = st.text_input("Password", type="password")

    if st.button("Log In"):
        user = cursor.execute("SELECT username, role FROM users WHERE username = ? AND password = ?", 
                              (input_username.strip(), input_password.strip())).fetchone()
        if user:
            st.session_state["logged_in"] = True
            st.session_state["username"] = user[0]
            st.session_state["user_role"] = user[1]
            st.success(f"Welcome back, {user[0]}!")
            st.rerun()
        else:
            st.error("Invalid username or password.")

# --- MAIN APPLICATION (LOGGED IN) ---
else:
    user_name = st.session_state["username"]
    user_role = st.session_state["user_role"]

    unread_query = """
        SELECT COUNT(*) FROM notifications 
        WHERE is_read = 0 AND (target_user = ? OR target_role = ?)
    """
    unread_count = cursor.execute(unread_query, (user_name, user_role)).fetchone()[0]

    bell_label = f"🔔 ({unread_count})" if unread_count > 0 else "🔔"

    header_col1, header_col2 = st.columns([0.8, 0.2])
    with header_col1:
        st.title("🏗️ Construction Site Inventory System")
    with header_col2:
        st.space()
        with st.popover(bell_label, use_container_width=True):
            st.markdown(f"### Notifications for `{user_name}`")
            
            notifs = cursor.execute("""
                SELECT id, message, timestamp, is_read FROM notifications 
                WHERE target_user = ? OR target_role = ?
                ORDER BY id DESC LIMIT 15
            """, (user_name, user_role)).fetchall()

            if notifs:
                if st.button("Mark All as Read", key="read_all_notifs"):
                    cursor.execute("""
                        UPDATE notifications SET is_read = 1 
                        WHERE target_user = ? OR target_role = ?
                    """, (user_name, user_role))
                    conn.commit()
                    st.rerun()

                st.markdown("---")
                for n_id, msg, ts, is_r in notifs:
                    badge = "🟢 **NEW** " if is_r == 0 else ""
                    st.markdown(f"{badge}*{ts}*\n\n{msg}")
                    st.divider()
            else:
                st.info("No notifications yet.")

    # --- SIDEBAR NAVIGATION & USER INFO ---
    st.sidebar.markdown(f"**Logged in as:** `{user_name}`")
    st.sidebar.markdown(f"**Role:** `{user_role}`")
    st.sidebar.markdown("---")

    # Navigation options updated with Physical Inventory module
    if user_role == "Materials Supervisor":
        nav_options = [
            "📋 Current Inventory", 
            "📊 Analytics",
            "+ Stock In", 
            "- Stock Out",
            "📦 Physical Inventory Audit",
            "📜 My Log & Request Edits",
            "📅 Daily Report (Excel)",
            "⏰ Reminders",
            "📅 Schedule"
        ]
    else:  # Head Office Admin
        pending_requests_count = cursor.execute("SELECT COUNT(*) FROM edit_requests WHERE status = 'PENDING'").fetchone()[0]
        edit_option_title = f"✏️ Edit Requests ({pending_requests_count})" if pending_requests_count > 0 else "✏️ Edit Requests"

        nav_options = [
            "📋 Current Inventory", 
            "📊 Analytics",
            "+ Stock In", 
            "- Stock Out", 
            "📦 Physical Inventory Audit",
            edit_option_title,
            "➕ Manage Master Items", 
            "📜 Master Audit Log",
            "👤 Manage Users",
            "⏰ Reminders",
            "📅 Schedule"
        ]

    st.sidebar.markdown("### 📌 Navigation")
    selected_menu = st.sidebar.radio("Go to:", nav_options)

    st.sidebar.markdown("---")
    if st.sidebar.button("Log Out", use_container_width=True):
        st.session_state["logged_in"] = False
        st.session_state["username"] = None
        st.session_state["user_role"] = None
        st.rerun()

    # --- ROUTING BASED ON SIDEBAR SELECTION ---

    # --- MENU 1: CURRENT INVENTORY ---
    if selected_menu == "📋 Current Inventory":
        st.subheader("📋 Current Available Stocks")

        df_items = pd.read_sql_query(
            "SELECT category, item_name, current_stock, min_threshold, unit FROM master_items ORDER BY category ASC, item_name ASC", 
            conn
        )
        
        if not df_items.empty:
            def highlight_low_stock(row):
                return ['background-color: #ffcccc' if row['current_stock'] <= row['min_threshold'] else '' for _ in row]

            categories = df_items['category'].unique()
            for cat in categories:
                st.markdown(f"### 📂 **{cat.upper()}**")
                
                df_cat = df_items[df_items['category'] == cat][['item_name', 'current_stock', 'min_threshold', 'unit']]

                st.dataframe(
                    df_cat.style.apply(highlight_low_stock, axis=1),
                    column_config={
                        "item_name": st.column_config.Column("Item Name", pinned=True, width="medium"),
                        "current_stock": st.column_config.NumberColumn("Current Stock", format="%.2f"),
                        "min_threshold": st.column_config.NumberColumn("Min. Threshold", format="%.2f"),
                        "unit": st.column_config.Column("Unit")
                    },
                    hide_index=True,
                    use_container_width=True
                )
        else:
            st.info("No items found in Master Inventory.")

    # --- MENU 2: ANALYTICS DASHBOARD ---
    elif selected_menu == "📊 Analytics":
        st.subheader("📊 Inventory Analytics & Operational Insights")

        filter_col1, filter_col2 = st.columns([0.4, 0.6])
        with filter_col1:
            time_view = st.selectbox(
                "🗓️ Select Time View:", 
                ["Daily (Last 14 Days)", "Weekly (Last 8 Weeks)", "Monthly (Last 12 Months)", "All Time / Year-To-Date"]
            )
        
        today = datetime.now()
        
        if time_view == "Daily (Last 14 Days)":
            start_date = today - timedelta(days=14)
            freq_rule = "D"
            date_format = "%Y-%m-%d"
        elif time_view == "Weekly (Last 8 Weeks)":
            start_date = today - timedelta(weeks=8)
            freq_rule = "W-MON"
            date_format = "Week %U, %Y"
        elif time_view == "Monthly (Last 12 Months)":
            start_date = today - timedelta(days=365)
            freq_rule = "ME"
            date_format = "%b %Y"
        else:
            start_date = datetime(today.year, 1, 1)
            freq_rule = "ME"
            date_format = "%b %Y"

        df_inv = pd.read_sql_query("SELECT item_name, category, unit, current_stock, min_threshold FROM master_items", conn)
        df_tx_all = pd.read_sql_query("SELECT timestamp, item_name, type, quantity, user_role, driver_details, project_name FROM transactions", conn)

        if df_inv.empty:
            st.warning("No inventory data available for analytics.")
        else:
            if not df_tx_all.empty:
                df_tx_all['dt_timestamp'] = pd.to_datetime(df_tx_all['timestamp'])
                df_filtered_tx = df_tx_all[df_tx_all['dt_timestamp'] >= start_date].copy()
            else:
                df_filtered_tx = pd.DataFrame()

            total_skus = len(df_inv)
            low_stock_items = df_inv[df_inv['current_stock'] <= df_inv['min_threshold']]
            low_stock_count = len(low_stock_items)

            total_received_qty = df_filtered_tx[df_filtered_tx['type'] == 'IN']['quantity'].sum() if not df_filtered_tx.empty else 0
            total_issued_qty = df_filtered_tx[df_filtered_tx['type'] == 'OUT']['quantity'].sum() if not df_filtered_tx.empty else 0

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Active SKUs", total_skus)
            m2.metric("Low Stock Alerts", low_stock_count, delta=f"{low_stock_count} Critical", delta_color="inverse" if low_stock_count > 0 else "off")
            m3.metric(f"Received (IN) [{time_view.split()[0]}]", f"{total_received_qty:,.1f}")
            m4.metric(f"Issued (OUT) [{time_view.split()[0]}]", f"{total_issued_qty:,.1f}")

            st.markdown("---")

            if low_stock_count > 0:
                st.error(f"⚠️ **Attention Required:** {low_stock_count} item(s) are at or below safety threshold!")
                st.dataframe(low_stock_items[['item_name', 'category', 'current_stock', 'min_threshold', 'unit']], use_container_width=True)
            else:
                st.success("✅ Stock levels across all items are healthy.")

            st.markdown("---")

            st.markdown(f"##### 📈 Stock Movement Trends ({time_view})")
            if not df_filtered_tx.empty:
                df_filtered_tx_grouped = df_filtered_tx.set_index('dt_timestamp')
                trend_in = df_filtered_tx_grouped[df_filtered_tx_grouped['type'] == 'IN'].resample(freq_rule)['quantity'].sum()
                trend_out = df_filtered_tx_grouped[df_filtered_tx_grouped['type'] == 'OUT'].resample(freq_rule)['quantity'].sum()

                df_trend = pd.DataFrame({'Stock IN': trend_in, 'Stock OUT': trend_out}).fillna(0)
                df_trend.index = df_trend.index.strftime(date_format)

                st.line_chart(df_trend, use_container_width=True)
            else:
                st.info(f"No transactions recorded for the selected time window ({time_view}).")

            st.markdown("---")

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### 📦 Current Stock vs. Threshold (Top 15)")
                df_chart_inv = df_inv.sort_values(by='current_stock', ascending=False).head(15)
                st.bar_chart(df_chart_inv, x="item_name", y=["current_stock", "min_threshold"], use_container_width=True)

            with c2:
                st.markdown("##### 🏷️ Category-wise SKU Count")
                cat_summary = df_inv.groupby("category")["item_name"].count().reset_index()
                cat_summary.columns = ["Category", "Total Items"]
                st.bar_chart(cat_summary, x="Category", y="Total Items", use_container_width=True)

            st.markdown("---")

            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown(f"##### 🔥 Top Issued Items ({time_view.split()[0]})")
                if not df_filtered_tx.empty and not df_filtered_tx[df_filtered_tx['type'] == 'OUT'].empty:
                    df_out_top = df_filtered_tx[df_filtered_tx['type'] == 'OUT'].groupby('item_name')['quantity'].sum().reset_index()
                    df_out_top = df_out_top.sort_values(by='quantity', ascending=False).head(10)
                    st.dataframe(df_out_top, use_container_width=True)
                else:
                    st.caption("No Stock OUT transactions logged in this period.")

            with col_b:
                st.markdown(f"##### 🚚 Active Drivers & Transport Logins ({time_view.split()[0]})")
                if not df_filtered_tx.empty and df_filtered_tx['driver_details'].notnull().any():
                    df_drivers = df_filtered_tx[df_filtered_tx['driver_details'] != ''].groupby('driver_details')['timestamp'].count().reset_index()
                    df_drivers.columns = ['Driver / Vehicle Details', 'Trips Logged']
                    df_drivers = df_drivers.sort_values(by='Trips Logged', ascending=False).head(10)
                    st.dataframe(df_drivers, use_container_width=True)
                else:
                    st.caption("No driver entries logged in this period.")

    # --- MENU 3: STOCK IN ---
    elif selected_menu == "+ Stock In":
        st.subheader("Log Stock Delivery (Receiving)")
        items = [row[0] for row in cursor.execute("SELECT item_name FROM master_items").fetchall()]
        if items:
            col_a, col_b = st.columns(2)
            with col_a:
                item_selected = st.selectbox("Select Item to Receive", items, key="in_item")
                qty_in = st.number_input("Quantity Received", min_value=0.1, step=1.0, key="in_qty")
            with col_b:
                driver_info = st.text_input("Driver / Delivery Details / DR #", placeholder="e.g. John Doe (Plate: ABC-123, DR #9876)", key="in_driver")
                remarks_in = st.text_input("General Remarks", placeholder="e.g. Verified good condition, unloaded at Bay 2", key="in_rem")

            st.markdown("### 📷 Delivery Photo / Receipt Proof")
            photo_mode = st.radio("Choose Photo Upload Method:", ["Camera Capture", "Upload File"], horizontal=True, key="in_photo_mode")
            
            image_bytes = None
            if photo_mode == "Camera Capture":
                camera_photo = st.camera_input("Take a picture of the delivery / receipt", key="in_camera")
                if camera_photo:
                    image_bytes = camera_photo.getvalue()
            else:
                uploaded_file = st.file_uploader("Upload Delivery Receipt / Photo", type=["jpg", "jpeg", "png"], key="in_upload")
                if uploaded_file:
                    image_bytes = uploaded_file.getvalue()

            if st.button("Submit Stock In", use_container_width=True):
                saved_photo_path = None
                if image_bytes:
                    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    file_filename = f"uploads/delivery_{timestamp_str}.jpg"
                    with open(file_filename, "wb") as f:
                        f.write(image_bytes)
                    saved_photo_path = file_filename

                cursor.execute("UPDATE master_items SET current_stock = current_stock + ? WHERE item_name = ?", (qty_in, item_selected))
                cursor.execute("""
                    INSERT INTO transactions (timestamp, item_name, type, quantity, user_role, driver_details, remarks, photo_path) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), item_selected, "IN", qty_in, user_name, driver_info, remarks_in, saved_photo_path))
                conn.commit()
                st.success(f"Successfully recorded receiving {qty_in} of {item_selected}!")
                st.rerun()
        else:
            st.warning("Please add items to Master Inventory first.")

    # --- MENU 4: STOCK OUT ---
    elif selected_menu == "- Stock Out":
        st.subheader("Log Stock Issuance (Stock OUT)")
        items = [row[0] for row in cursor.execute("SELECT item_name FROM master_items").fetchall()]
        if items:
            col_x, col_y = st.columns(2)
            with col_x:
                item_selected = st.selectbox("Select Item to Issue", items, key="out_item")
                qty_out = st.number_input("Quantity Issued", min_value=0.1, step=1.0, key="out_qty")
                issued_to = st.text_input("Issued To (Person / Subcontractor)", placeholder="e.g. Foreman Mike / Subcon ABC", key="out_issued_to")
            
            with col_y:
                driver_out = st.text_input("Driver / Transport Vehicle", placeholder="e.g. Driver Bob (Dump Truck #2)", key="out_driver")
                project_name = st.text_input("Project Name / Site Location", placeholder="e.g. Tower B - 5th Floor", key="out_project")
                purpose = st.text_input("Purpose / Equipment Usage", placeholder="e.g. Concrete Pouring Foundation", key="out_purpose")

            current_stk, min_thresh = cursor.execute("SELECT current_stock, min_threshold FROM master_items WHERE item_name = ?", (item_selected,)).fetchone()
            st.caption(f"Current Stock Available: **{current_stk}** | Min Alert Threshold: **{min_thresh}**")

            st.markdown("### 📷 Requisition / Issuance Proof Photo")
            photo_mode_out = st.radio("Choose Photo Upload Method:", ["Camera Capture", "Upload File"], horizontal=True, key="out_photo_mode")
            
            image_bytes_out = None
            if photo_mode_out == "Camera Capture":
                camera_photo_out = st.camera_input("Take a picture of the requisition form / materials", key="out_camera")
                if camera_photo_out:
                    image_bytes_out = camera_photo_out.getvalue()
            else:
                uploaded_file_out = st.file_uploader("Upload Requisition Slip / Photo", type=["jpg", "jpeg", "png"], key="out_upload")
                if uploaded_file_out:
                    image_bytes_out = uploaded_file_out.getvalue()

            if st.button("Submit Stock Out", use_container_width=True):
                if qty_out > current_stk:
                    st.error(f"Cannot issue {qty_out}. Stock available is only {current_stk}!")
                else:
                    saved_photo_path = None
                    if image_bytes_out:
                        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                        file_filename = f"uploads/issuance_{timestamp_str}.jpg"
                        with open(file_filename, "wb") as f:
                            f.write(image_bytes_out)
                        saved_photo_path = file_filename

                    new_stock = current_stk - qty_out
                    cursor.execute("UPDATE master_items SET current_stock = ? WHERE item_name = ?", (new_stock, item_selected))
                    cursor.execute("""
                        INSERT INTO transactions (timestamp, item_name, type, quantity, user_role, driver_details, issued_to, project_name, purpose, photo_path) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), item_selected, "OUT", qty_out, user_name, driver_out, issued_to, project_name, purpose, saved_photo_path))
                    conn.commit()

                    if new_stock <= min_thresh:
                        create_notification(f"⚠️ LOW STOCK ALERT: {item_selected} stock dropped to {new_stock} (Threshold: {min_thresh})", target_role="Head Office")

                    st.success(f"Successfully issued {qty_out} of {item_selected}!")
                    st.rerun()
        else:
            st.warning("Please add items to Master Inventory first.")

    # --- MENU: PHYSICAL INVENTORY AUDIT (NEW FEATURE) ---
    elif selected_menu == "📦 Physical Inventory Audit":
        st.subheader("📦 Physical Inventory Count & Variance Audit")

        tab_audit, tab_history = st.tabs(["📝 Record Physical Count", "📜 Audit History & Discrepancies"])

        with tab_audit:
            items_master = cursor.execute("SELECT item_name, current_stock, unit FROM master_items ORDER BY item_name ASC").fetchall()
            
            if items_master:
                item_dict = {row[0]: {"sys_stock": row[1], "unit": row[2]} for row in items_master}
                
                col1, col2 = st.columns(2)
                with col1:
                    audit_date = st.date_input("Audit Date", datetime.now(), key="phys_date")
                    selected_item = st.selectbox("Select Item to Audit", list(item_dict.keys()), key="phys_item")
                
                with col2:
                    system_qty = item_dict[selected_item]["sys_stock"]
                    item_unit = item_dict[selected_item]["unit"]
                    
                    st.metric(f"Recorded System Stock ({item_unit})", f"{system_qty:,.2f}")
                    physical_qty = st.number_input(f"Actual Physical Count ({item_unit})", min_value=0.0, step=1.0, value=float(system_qty), key="phys_qty")

                variance = physical_qty - system_qty
                
                if variance == 0:
                    st.success("✅ Perfect Match: Physical count matches system record.")
                elif variance < 0:
                    st.error(f"⚠️ Stock Deficit Identified: Physical stock is **{abs(variance):,.2f} {item_unit} LESS** than system stock.")
                else:
                    st.warning(f"ℹ️ Stock Surplus Identified: Physical stock is **{variance:,.2f} {item_unit} MORE** than system stock.")

                audit_remarks = st.text_area("Audit Notes / Explanation for Variance", placeholder="e.g. Damaged goods removed, unrecorded issuance, or supplier over-delivery")

                if st.button("Submit Physical Count", use_container_width=True):
                    cursor.execute("""
                        INSERT INTO physical_inventory_counts 
                        (audit_date, item_name, system_stock, physical_stock, variance, unit, counted_by, remarks, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (audit_date.strftime("%Y-%m-%d"), selected_item, system_qty, physical_qty, variance, item_unit, user_name, audit_remarks, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    
                    conn.commit()

                    if variance != 0:
                        create_notification(
                            f"📦 Physical count discrepancy logged for '{selected_item}': Variance = {variance:+,.2f} {item_unit} (Audited by {user_name})", 
                            target_role="Head Office"
                        )

                    st.success(f"Physical count audit logged for {selected_item}!")
                    st.rerun()

        with tab_history:
            st.markdown("### Physical Inventory Audit Log")
            df_audits = pd.read_sql_query("SELECT * FROM physical_inventory_counts ORDER BY id DESC", conn)

            if not df_audits.empty:
                def highlight_variance(val):
                    if val < 0:
                        return 'color: red; font-weight: bold;'
                    elif val > 0:
                        return 'color: orange; font-weight: bold;'
                    return 'color: green;'

                st.dataframe(
                    df_audits.style.applymap(highlight_variance, subset=['variance']),
                    use_container_width=True
                )

                st.markdown("---")
                st.markdown("### 🔄 Sync System Inventory to Physical Count")
                st.caption("Reconcile system stock levels directly to match verified physical counts.")

                unsynced_audits = df_audits[df_audits['sync_status'] == 'UNSYNCED']

                if not unsynced_audits.empty:
                    sync_target_id = st.selectbox(
                        "Select Audit Entry to Sync:", 
                        unsynced_audits['id'].tolist(),
                        format_func=lambda x: f"Audit #{x} - {unsynced_audits[unsynced_audits['id']==x]['item_name'].values[0]} (Variance: {unsynced_audits[unsynced_audits['id']==x]['variance'].values[0]:+,.2f})"
                    )

                    target_row = unsynced_audits[unsynced_audits['id'] == sync_target_id].iloc[0]

                    if st.button(f"Reconcile Stock for {target_row['item_name']} to {target_row['physical_stock']} {target_row['unit']}"):
                        cursor.execute("UPDATE master_items SET current_stock = ? WHERE item_name = ?", (target_row['physical_stock'], target_row['item_name']))
                        cursor.execute("UPDATE physical_inventory_counts SET sync_status = 'SYNCED' WHERE id = ?", (sync_target_id,))
                        
                        cursor.execute("""
                            INSERT INTO transactions (timestamp, item_name, type, quantity, user_role, remarks) 
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                            target_row['item_name'], 
                            "AUDIT_ADJUSTMENT", 
                            target_row['variance'], 
                            user_name, 
                            f"Stock adjusted via Physical Audit #{sync_target_id} (Reason: {target_row['remarks']})"
                        ))

                        conn.commit()
                        st.success(f"System stock for '{target_row['item_name']}' reconciled to {target_row['physical_stock']}!")
                        st.rerun()
                else:
                    st.info("All physical audit entries are synced with system inventory.")
            else:
                st.info("No physical inventory count records logged yet.")

    # --- MENU 5 (SUPERVISOR): MY LOG & REQUEST EDITS ---
    elif selected_menu == "📜 My Log & Request Edits":
        st.subheader("📜 My Transaction Log & Edit Requests")
        
        my_txs = pd.read_sql_query(
            "SELECT * FROM transactions WHERE user_role = ? ORDER BY id DESC", 
            conn, 
            params=(user_name,)
        )

        if not my_txs.empty:
            st.dataframe(my_txs, use_container_width=True)

            st.markdown("---")
            st.markdown("### ✏️ Request Transaction Correction")
            
            selected_tx_id = st.selectbox("Select Transaction ID to Edit:", my_txs['id'].tolist())
            tx_row = my_txs[my_txs['id'] == selected_tx_id].iloc[0]

            with st.form("edit_request_form"):
                st.write(f"**Modifying Transaction ID:** `{selected_tx_id}` ({tx_row['item_name']} - {tx_row['type']})")
                new_qty = st.number_input("Corrected Quantity", value=float(tx_row['quantity']))
                new_remarks = st.text_input("Updated Purpose / Remarks", value=str(tx_row['remarks'] or tx_row['purpose'] or ""))
                reason = st.text_area("Reason for Edit Request (Required)")

                submit_req = st.form_submit_button("Submit Edit Request")

                if submit_req:
                    if not reason.strip():
                        st.error("Please provide a reason for requesting this edit.")
                    else:
                        orig_data = json.dumps(tx_row.to_dict(), default=str)
                        prop_dict = tx_row.to_dict()
                        prop_dict['quantity'] = new_qty
                        prop_dict['remarks'] = new_remarks
                        prop_data = json.dumps(prop_dict, default=str)

                        cursor.execute("""
                            INSERT INTO edit_requests (transaction_id, requested_by, reason, original_data, proposed_data, timestamp)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (selected_tx_id, user_name, reason, orig_data, prop_data, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

                        cursor.execute("UPDATE transactions SET edit_status = 'PENDING_EDIT' WHERE id = ?", (selected_tx_id,))
                        conn.commit()

                        create_notification(f"✏️ New edit request submitted by {user_name} for TX #{selected_tx_id}", target_role="Head Office")
                        st.success("Edit request submitted to Head Office for approval!")
                        st.rerun()
        else:
            st.info("You have not logged any transactions yet.")

    # --- MENU 5 (ADMIN): EDIT REQUESTS ---
    elif selected_menu.startswith("✏️ Edit Requests"):
        st.subheader("✏️ Pending Edit Requests")

        requests_df = pd.read_sql_query("SELECT * FROM edit_requests WHERE status = 'PENDING' ORDER BY id DESC", conn)

        if not requests_df.empty:
            for _, r in requests_df.iterrows():
                with st.expander(f"Request ID #{r['id']} - Transaction ID #{r['transaction_id']} (by {r['requested_by']})"):
                    orig = json.loads(r['original_data'])
                    prop = json.loads(r['proposed_data'])

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**Original Data:**")
                        st.json(orig)
                    with col2:
                        st.markdown("**Proposed Data:**")
                        st.json(prop)

                    st.markdown(f"**Reason for Request:** {r['reason']}")
                    review_remark = st.text_input("Review Remarks (Optional)", key=f"rev_{r['id']}")

                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("Approve Request", key=f"app_{r['id']}"):
                            item_name = orig['item_name']
                            tx_type = orig['type']
                            old_qty = float(orig['quantity'])
                            new_qty = float(prop['quantity'])
                            qty_diff = new_qty - old_qty

                            if tx_type == "IN":
                                cursor.execute("UPDATE master_items SET current_stock = current_stock + ? WHERE item_name = ?", (qty_diff, item_name))
                            else:
                                cursor.execute("UPDATE master_items SET current_stock = current_stock - ? WHERE item_name = ?", (qty_diff, item_name))

                            cursor.execute("""
                                UPDATE transactions 
                                SET quantity = ?, remarks = ?, edit_status = 'EDITED' 
                                WHERE id = ?
                            """, (new_qty, prop.get('remarks', orig.get('remarks')), r['transaction_id']))

                            cursor.execute("""
                                UPDATE edit_requests 
                                SET status = 'APPROVED', review_remarks = ? 
                                WHERE id = ?
                            """, (review_remark, r['id']))

                            conn.commit()
                            create_notification(f"✅ Your edit request for TX #{r['transaction_id']} was APPROVED.", target_user=r['requested_by'])
                            st.success("Request approved and inventory levels adjusted!")
                            st.rerun()

                    with btn_col2:
                        if st.button("Reject Request", key=f"rej_{r['id']}"):
                            cursor.execute("UPDATE transactions SET edit_status = 'NORMAL' WHERE id = ?", (r['transaction_id'],))
                            cursor.execute("""
                                UPDATE edit_requests 
                                SET status = 'REJECTED', review_remarks = ? 
                                WHERE id = ?
                            """, (review_remark, r['id']))

                            conn.commit()
                            create_notification(f"❌ Your edit request for TX #{r['transaction_id']} was REJECTED.", target_user=r['requested_by'])
                            st.warning("Request rejected.")
                            st.rerun()
        else:
            st.info("No pending edit requests found.")

    # --- MENU 6 (SUPERVISOR): DAILY REPORT EXCEL ---
    elif selected_menu == "📅 Daily Report (Excel)":
        st.subheader("📅 Export Daily Inventory Report")

        selected_date = st.date_input("Select Report Date", datetime.now())
        date_str = selected_date.strftime("%Y-%m-%d")

        df_daily_tx = pd.read_sql_query("""
            SELECT * FROM transactions 
            WHERE DATE(timestamp) = DATE(?) 
            ORDER BY id ASC
        """, conn, params=(date_str,))

        df_current_stock = pd.read_sql_query("SELECT * FROM master_items ORDER BY category ASC, item_name ASC", conn)

        st.markdown(f"**Found {len(df_daily_tx)} transactions for {date_str}.**")
        
        excel_data = generate_excel_report(user_name, date_str, df_daily_tx, df_current_stock)

        st.download_button(
            label="📥 Download Multi-Sheet Excel Report",
            data=excel_data,
            file_name=f"Inventory_Report_{date_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    # --- MENU 6 (ADMIN): MANAGE MASTER ITEMS ---
    elif selected_menu == "➕ Manage Master Items":
        st.subheader("➕ Manage Master Inventory Catalog")

        tab1, tab2 = st.tabs(["Add New Item", "Update Existing Item Thresholds"])

        with tab1:
            with st.form("add_item_form"):
                new_item_name = st.text_input("Item Name")
                new_category = st.text_input("Category")
                new_unit = st.text_input("Unit of Measure (e.g. Bags, Pcs, Liters)")
                initial_stock = st.number_input("Initial Stock Level", min_value=0.0, step=1.0)
                min_threshold = st.number_input("Minimum Alert Threshold", min_value=0.0, step=1.0)

                submit_item = st.form_submit_button("Add Item to Catalog")

                if submit_item:
                    if new_item_name and new_category:
                        try:
                            cursor.execute("""
                                INSERT INTO master_items (item_name, category, unit, current_stock, min_threshold)
                                VALUES (?, ?, ?, ?, ?)
                            """, (new_item_name.strip(), new_category.strip(), new_unit.strip(), initial_stock, min_threshold))
                            conn.commit()
                            st.success(f"Item '{new_item_name}' added successfully!")
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error(f"Item '{new_item_name}' already exists in database.")
                    else:
                        st.error("Item Name and Category are required.")

        with tab2:
            items_df = pd.read_sql_query("SELECT * FROM master_items ORDER BY item_name ASC", conn)
            if not items_df.empty:
                selected_item = st.selectbox("Select Item to Update:", items_df['item_name'].tolist())
                item_data = items_df[items_df['item_name'] == selected_item].iloc[0]

                with st.form("update_item_form"):
                    up_cat = st.text_input("Category", value=item_data['category'])
                    up_unit = st.text_input("Unit", value=item_data['unit'])
                    up_thresh = st.number_input("Min Alert Threshold", value=float(item_data['min_threshold']))

                    submit_update = st.form_submit_button("Save Changes")

                    if submit_update:
                        cursor.execute("""
                            UPDATE master_items 
                            SET category = ?, unit = ?, min_threshold = ? 
                            WHERE item_name = ?
                        """, (up_cat.strip(), up_unit.strip(), up_thresh, selected_item))
                        conn.commit()
                        st.success(f"Updated details for {selected_item}!")
                        st.rerun()

    # --- MENU 7 (ADMIN): MASTER AUDIT LOG ---
    elif selected_menu == "📜 Master Audit Log":
        st.subheader("📜 Complete Site Transaction Audit Trail")

        df_audit = pd.read_sql_query("SELECT * FROM transactions ORDER BY id DESC", conn)
        if not df_audit.empty:
            st.dataframe(df_audit, use_container_width=True)
        else:
            st.info("No transaction records found.")

    # --- MENU 8 (ADMIN): MANAGE USERS ---
    elif selected_menu == "👤 Manage Users":
        st.subheader("👤 User Account Management")

        with st.form("create_user_form"):
            st.markdown("### Create New User")
            new_u = st.text_input("Username")
            new_p = st.text_input("Password", type="password")
            new_r = st.selectbox("Role", ["Materials Supervisor", "Head Office"])

            create_u = st.form_submit_button("Create Account")

            if create_u:
                if new_u and new_p:
                    try:
                        cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", (new_u.strip(), new_p.strip(), new_r))
                        conn.commit()
                        st.success(f"Account for '{new_u}' created!")
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("Username already taken.")
                else:
                    st.error("Please fill in both username and password.")

        st.markdown("---")
        st.markdown("### Existing Users")
        users_df = pd.read_sql_query("SELECT id, username, role FROM users", conn)
        st.dataframe(users_df, use_container_width=True)

    # --- MENU: REMINDERS ---
    elif selected_menu == "⏰ Reminders":
        st.subheader("⏰ Personal & Task Reminders")

        with st.form("add_reminder_form"):
            r_title = st.text_input("Reminder Description / Task")
            r_due = st.date_input("Due Date", datetime.now())
            r_prio = st.selectbox("Priority Level", ["Low", "Medium", "High"])

            sub_rem = st.form_submit_button("Set Reminder")

            if sub_rem:
                if r_title:
                    cursor.execute("""
                        INSERT INTO reminders (user_name, title, due_date, priority, timestamp)
                        VALUES (?, ?, ?, ?, ?)
                    """, (user_name, r_title.strip(), r_due.strftime("%Y-%m-%d"), r_prio, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    conn.commit()
                    st.success("Reminder created!")
                    st.rerun()

        st.markdown("---")
        st.markdown("### Active Tasks & Reminders")
        rems = pd.read_sql_query("SELECT id, title, due_date, priority, status FROM reminders WHERE user_name = ? ORDER BY due_date ASC", conn, params=(user_name,))

        if not rems.empty:
            for _, r in rems.iterrows():
                col1, col2, col3 = st.columns([0.6, 0.2, 0.2])
                with col1:
                    status_flag = "✅ " if r['status'] == 'COMPLETED' else "⏳ "
                    st.markdown(f"{status_flag}**{r['title']}** (Due: {r['due_date']}) - *Priority: {r['priority']}*")
                with col2:
                    if r['status'] != 'COMPLETED':
                        if st.button("Mark Done", key=f"done_rem_{r['id']}"):
                            cursor.execute("UPDATE reminders SET status = 'COMPLETED' WHERE id = ?", (r['id'],))
                            conn.commit()
                            st.rerun()
                with col3:
                    if st.button("Delete", key=f"del_rem_{r['id']}"):
                        cursor.execute("DELETE FROM reminders WHERE id = ?", (r['id'],))
                        conn.commit()
                        st.rerun()
        else:
            st.info("No active reminders.")

    # --- MENU: SCHEDULE ---
    elif selected_menu == "📅 Schedule":
        st.subheader("📅 Site Events & Delivery Schedules")

        with st.form("add_schedule_form"):
            s_title = st.text_input("Event Title / Delivery Notice")
            s_date = st.date_input("Event Date", datetime.now())
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                s_start = st.time_input("Start Time", time(8, 0))
            with col_t2:
                s_end = st.time_input("End Time", time(17, 0))
            s_loc = st.text_input("Location / Gate / Subcon")
            s_notes = st.text_area("Notes / Additional Information")

            sub_sched = st.form_submit_button("Add Event to Schedule")

            if sub_sched:
                if s_title:
                    cursor.execute("""
                        INSERT INTO schedules (user_name, title, event_date, start_time, end_time, location_details, notes, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (user_name, s_title.strip(), s_date.strftime("%Y-%m-%d"), s_start.strftime("%H:%M"), s_end.strftime("%H:%M"), s_loc.strip(), s_notes.strip(), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    conn.commit()
                    st.success("Event scheduled!")
                    st.rerun()

        st.markdown("---")
        st.markdown("### Upcoming Events")
        scheds = pd.read_sql_query("SELECT id, title, event_date, start_time, end_time, location_details, notes FROM schedules ORDER BY event_date ASC, start_time ASC", conn)

        if not scheds.empty:
            for _, s in scheds.iterrows():
                with st.expander(f"📅 {s['event_date']} | {s['start_time']} - {s['end_time']} : {s['title']}"):
                    st.write(f"**Location / Vehicle:** {s['location_details']}")
                    st.write(f"**Notes:** {s['notes']}")
                    if st.button("Delete Event", key=f"del_sch_{s['id']}"):
                        cursor.execute("DELETE FROM schedules WHERE id = ?", (s['id'],))
                        conn.commit()
                        st.rerun()
        else:
            st.info("No scheduled events available.")
